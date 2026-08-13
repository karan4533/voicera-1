"""Simple manager-friendly Voicera security status sheet."""
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BROWN = "50381F"
CREAM = "F7F4EF"
SURFACE = "ECE6D9"
WHITE = "FFFFFF"
TEXT = "1E1A16"
MUTED = "6B645B"
GREEN_BG = "E8F5E9"
AMBER_BG = "FFF8E1"
RED_BG = "FFEBEE"

header_fill = PatternFill("solid", fgColor=BROWN)
header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
title_font = Font(name="Calibri", bold=True, color=BROWN, size=20)
subtitle_font = Font(name="Calibri", italic=True, color=MUTED, size=11)
label_font = Font(name="Calibri", bold=True, color=TEXT, size=11)
body_font = Font(name="Calibri", color=TEXT, size=11)
done_fill = PatternFill("solid", fgColor=GREEN_BG)
prog_fill = PatternFill("solid", fgColor=AMBER_BG)
open_fill = PatternFill("solid", fgColor=RED_BG)
cream_fill = PatternFill("solid", fgColor=CREAM)
white_fill = PatternFill("solid", fgColor=WHITE)
surface_fill = PatternFill("solid", fgColor=SURFACE)
thin = Border(
    left=Side(style="thin", color="E7DFC8"),
    right=Side(style="thin", color="E7DFC8"),
    top=Side(style="thin", color="E7DFC8"),
    bottom=Side(style="thin", color="E7DFC8"),
)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")


def header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def paint_status(cell):
    v = str(cell.value or "").lower()
    if v == "done":
        cell.fill = done_fill
    elif v in ("in progress", "partial"):
        cell.fill = prog_fill
    else:
        cell.fill = open_fill
    cell.alignment = center
    cell.font = Font(name="Calibri", bold=True, color=TEXT, size=11)


def body_row(ws, r, cols):
    ws.row_dimensions[r].height = 48
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin
        if c != 3:
            cell.fill = cream_fill if r % 2 == 0 else white_fill


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Status"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Voicera — Production Security Status"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Simple status for manager review  |  {date.today().strftime('%d %b %Y')}  |  Confidential — Internal"
    )
    ws["A2"].font = subtitle_font

    ws.merge_cells("A3:E3")
    ws["A3"] = (
        "In one line: core security (who can see whose data) is now in the code. "
        "It still needs to be deployed, then MFA and audit log can follow."
    )
    ws["A3"].font = label_font
    ws["A3"].fill = surface_fill
    ws["A3"].alignment = wrap
    ws.row_dimensions[3].height = 36

    # Main simple table
    headers = ["#", "Security topic", "Status", "What this means (plain English)", "What to do next"]
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)
    header_row(ws, 5, 5)
    ws.row_dimensions[5].height = 28

    rows = [
        [
            1,
            "Login / who is the user",
            "Done",
            "Users sign in with Firebase. After login they get a token. Admin goes to Admin Console, customer goes to their dashboard.",
            "Keep using Firebase login. Turn on MFA for admin accounts later.",
        ],
        [
            2,
            "One customer cannot see another customer's data",
            "Done",
            "Each company (tenant) can only read their own record. Only Platform Admin can see the full customer list or change accounts.",
            "Deploy Firestore rules to production so this is live.",
        ],
        [
            3,
            "Who can create a new customer account",
            "Done",
            "Only Platform Admin, and only through the server (Cloud Function). The browser can no longer create users by itself.",
            "Deploy Cloud Functions to production.",
        ],
        [
            4,
            "Passwords & secrets not stored in GitHub",
            "Done",
            "Env files and Firebase key files are ignored by git. Example env file has empty placeholders only.",
            "Confirm Vercel has real keys in dashboard env, not in the repo.",
        ],
        [
            5,
            "Website security headers (HTTPS, clickjacking, etc.)",
            "Done",
            "Vercel now sends extra browser protections (HSTS, CSP, no iframe embed).",
            "Redeploy the frontend on Vercel.",
        ],
        [
            6,
            "Mock / dummy data in production",
            "Done",
            "Production build will not show fake demo data even if someone forgets the env flag.",
            "Set VITE_USE_MOCK=false in Vercel environment.",
        ],
        [
            7,
            "Strong password for new customers",
            "Done",
            "New account password must be at least 12 characters (UI + server).",
            "No extra action after deploy.",
        ],
        [
            8,
            "Admin extra login (MFA)",
            "Open",
            "Brief asks that admin/owner must use MFA. This is not turned on yet in Firebase.",
            "Enable MFA for platform admin emails in Firebase Console.",
        ],
        [
            9,
            "Audit log (who did what, when)",
            "Open",
            "Enterprise customers always ask for this. Voicera does not store an activity log yet.",
            "Plan a simple audit table: user, action, time, company.",
        ],
        [
            10,
            "Rate limit / bot protection",
            "Open",
            "Brief asks to limit how many requests one IP can make. Not added yet.",
            "Turn on Firebase App Check (and WAF later).",
        ],
        [
            11,
            "Backup & incident plan",
            "Open",
            "No written backup/restore or incident response document for Voicera yet.",
            "Write a 1-page backup + incident note for the team.",
        ],
    ]

    for r, row in enumerate(rows, 6):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        body_row(ws, r, 5)
        paint_status(ws.cell(r, 3))
        ws.cell(r, 1).alignment = center

    # Summary box
    ws.merge_cells("A18:E18")
    ws["A18"] = "Quick summary"
    ws["A18"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws["A18"].fill = header_fill
    ws["A18"].alignment = center

    summary = [
        ["Done now (in code)", "Login, tenant isolation, admin-only account create, secrets out of git, security headers, no mock in prod, stronger password."],
        ["Must deploy this week", "1) Firestore rules  2) Cloud Functions  3) Vercel frontend  4) VITE_USE_MOCK=false"],
        ["Do next (after deploy)", "Admin MFA, App Check, audit log, backup/incident note."],
        ["Production ready?", "Not yet. Ready after the deploy steps above. Then we can say “production hardening started / core isolation live.”"],
    ]
    for i, (k, v) in enumerate(summary, 19):
        ws.cell(i, 1, k).font = label_font
        ws.cell(i, 1).fill = surface_fill
        ws.cell(i, 1).alignment = wrap
        ws.cell(i, 1).border = thin
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        ws.cell(i, 2, v).font = body_font
        ws.cell(i, 2).alignment = wrap
        ws.cell(i, 2).border = thin
        ws.row_dimensions[i].height = 40
        for c in range(2, 6):
            ws.cell(i, c).border = thin
            ws.cell(i, c).fill = white_fill if i % 2 else cream_fill

    ws.merge_cells("A24:E24")
    ws["A24"] = "Colour key:  Green = Done   |   Amber = In progress   |   Red = Open / not started"
    ws["A24"].font = subtitle_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 48
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = "A5:E16"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddFooter.center.text = "Voicera Security Status | Confidential — Internal | Page &P"

    # Sheet 2 — very small next-actions table
    ws2 = wb.create_sheet("Next actions")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Next actions (checklist)"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:D2")
    ws2["A2"] = "Update Status using the dropdown. This is the only follow-up sheet."
    ws2["A2"].font = subtitle_font

    h2 = ["When", "Action", "Who", "Status"]
    for i, h in enumerate(h2, 1):
        ws2.cell(4, i, h)
    header_row(ws2, 4, 4)

    actions = [
        ["This week", "Deploy Firestore rules to production", "DevOps / Eng", "Not Started"],
        ["This week", "Deploy Cloud Functions to production", "DevOps / Eng", "Not Started"],
        ["This week", "Redeploy frontend on Vercel", "DevOps / Eng", "Not Started"],
        ["This week", "Set VITE_USE_MOCK=false on Vercel", "DevOps", "Not Started"],
        ["This week", "Set platform_admin role on admin users", "Eng", "Not Started"],
        ["Next sprint", "Turn on MFA for admin accounts", "Eng + IT", "Not Started"],
        ["Next sprint", "Turn on Firebase App Check", "Eng", "Not Started"],
        ["Later", "Add audit log (who did what)", "Eng", "Not Started"],
        ["Later", "Write backup + incident 1-pager", "Eng + Manager", "Not Started"],
    ]
    for r, row in enumerate(actions, 5):
        for c, val in enumerate(row, 1):
            ws2.cell(r, c, val)
        body_row(ws2, r, 4)
        paint_status(ws2.cell(r, 4))
        ws2.row_dimensions[r].height = 28

    dv = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=False)
    ws2.add_data_validation(dv)
    dv.add("D5:D13")

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 16
    ws2.freeze_panes = "A5"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToWidth = 1
    ws2.oddFooter.center.text = "Voicera Security Status | Confidential — Internal | Page &P"

    out = r"c:\karan\voicera\Voicera_Security_Status_Simple.xlsx"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    build()
