"""Final frontend governance workbook from HL-SEC-DEVBRIEF-2026-001."""
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BROWN = "50381F"
CREAM = "F7F4EF"
SURFACE = "ECE6D9"
WHITE = "FFFFFF"
TEXT = "1E1A16"
MUTED = "6B645B"
GREEN = "E8F5E9"
AMBER = "FFF8E1"
RED = "FFEBEE"

header_fill = PatternFill("solid", fgColor=BROWN)
header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
title_font = Font(name="Calibri", bold=True, color=BROWN, size=18)
sub_font = Font(name="Calibri", italic=True, color=MUTED, size=11)
body = Font(name="Calibri", color=TEXT, size=11)
bold = Font(name="Calibri", bold=True, color=TEXT, size=11)
done_f = PatternFill("solid", fgColor=GREEN)
part_f = PatternFill("solid", fgColor=AMBER)
open_f = PatternFill("solid", fgColor=RED)
cream = PatternFill("solid", fgColor=CREAM)
white_f = PatternFill("solid", fgColor=WHITE)
surface = PatternFill("solid", fgColor=SURFACE)
thin = Border(
    left=Side(style="thin", color="E7DFC8"),
    right=Side(style="thin", color="E7DFC8"),
    top=Side(style="thin", color="E7DFC8"),
    bottom=Side(style="thin", color="E7DFC8"),
)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")


def status_fill(val: str):
    v = (val or "").lower()
    if v == "done":
        return done_f
    if v == "partial":
        return part_f
    return open_f


ROWS = [
    [1, "Login / Authentication (JWT)",
     "User signs in on the login page. Every API call sends the Firebase token. Do not store password in localStorage.",
     "Done",
     "Firebase Auth login. Token refresh is automatic. SSO/SAML not in frontend yet."],
    [2, "MFA (admin)",
     "Admin login should require extra verification (MFA) when Firebase MFA is enabled.",
     "Done",
     "TOTP enroll on Admin → Security. Login asks for authenticator code when MFA is on. Enable Identity Platform MFA in Firebase Console."],
    [3, "Tenant / User / Team management",
     "Show only the logged-in company workspace. Admin console only for platform admin. Do not let UI hide = security.",
     "Done",
     "RoleRoute: admin → /admin, customer → /dashboard. customer_user is view-only. Suspended account is blocked."],
    [4, "Data security / isolation (RLS)",
     "Frontend must request only own org data. Never list all customers from a customer login. Isolation must also be in Firestore rules.",
     "Done",
     "Rules: customer reads own org only. Admin can list all. Must deploy rules to go live."],
    [5, "Ownership check (no ID guess)",
     "Do not trust a URL or org id from the user. Server / Firestore rules decide if they own it.",
     "Done",
     "Create / update / offboard only via Cloud Function. No client-side user create."],
    [6, "Secrets never in frontend or git",
     "No API secret keys in React code. VITE_ Firebase keys are public by design. Real secrets stay in Vercel / Functions.",
     "Done",
     ".gitignore covers all .env files. .env.example has empty values."],
    [7, "HTTPS everywhere",
     "App must run on https in production. No http links for login or APIs.",
     "Done",
     "Vercel HTTPS + HSTS header."],
    [8, "Security headers (CSP, HSTS, X-Frame-Options)",
     "Frontend host must send browser protection headers so the app cannot be put in a fake iframe.",
     "Done",
     "Added in vercel.json. Needs Vercel redeploy."],
    [9, "Input validation / XSS",
     "Validate forms on screen (email, password length). Do not render untrusted HTML. Use React text, not innerHTML.",
     "Done",
     "Email + password policy on login/create. Names sanitised. React text only (no innerHTML)."],
    [10, "Session management",
     "Logout must clear session. Remember-me uses local storage; otherwise session only. Token expires and refreshes.",
     "Done",
     "Logout clears storage + revokes refresh tokens. Security page can sign out other devices."],
    [11, "Audit logging",
     "Frontend should not be the source of truth. Sensitive actions (create account, suspend) should be logged on server.",
     "Done",
     "Cloud Function stamps audit_events. Admin Security page lists last 50. Client cannot write the collection."],
    [12, "Logging hygiene / no PII in logs",
     "Do not console.log passwords, tokens, or full customer PII.",
     "Done",
     "safeLog redacts emails/secrets. Admin pages use it instead of raw console.error."],
    [13, "Credits / usage analytics UI",
     "If we show usage, show only that company's numbers. No cross-company charts for customers.",
     "Done",
     "Customer Usage & credits page shows own org calls vs plan limit only."],
    [14, "Rate limiting",
     "Frontend cannot fully enforce this. Still: disable double-submit on login/create buttons.",
     "Done",
     "Client throttle on login / MFA / reset / create-account. Loading disables double-submit. IP limit remains backend."],
    [15, "Dependency / supply-chain (frontend)",
     "Keep npm packages updated. Do not add unknown UI libraries without review.",
     "Done",
     "Dependabot weekly for root + functions. Review PRs before merge."],
    [16, "Data deletion / offboarding UI",
     "Admin should have a clear way to close/delete a customer when policy requires it.",
     "Done",
     "Admin drawer: Offboard / delete account. Cloud Function deletes Auth user + org + audit."],
    [17, "In-app notifications",
     "Header bell must show real events, not a dummy icon. Customer sees only own-org items.",
     "Done",
     "Live audit_events dropdown on admin + customer headers. Unread badge. Usage warning at 80%+."],
]


def add_table_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Frontend Security Table"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Voicera Frontend — Security & Compliance (FINAL)"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Based on HL-SEC-DEVBRIEF-2026-001  |  Frontend only  |  {date.today().strftime('%d %b %Y')}  |  Confidential — Internal"
    )
    ws["A2"].font = sub_font

    ws.merge_cells("A3:E3")
    ws["A3"] = (
        "Brief stack is Clerk / Supabase / FastAPI / Azure. Voicera frontend uses React + Firebase Auth + Firestore + Vercel. "
        "All 17 frontend items below are Done in code. Production sign-off still needs deploy."
    )
    ws["A3"].font = bold
    ws["A3"].fill = surface
    ws["A3"].alignment = wrap
    ws.row_dimensions[3].height = 40

    headers = ["#", "Governance topic (from brief)", "What frontend must do", "Status", "Simple note"]
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)
        ws.cell(5, i).fill = header_fill
        ws.cell(5, i).font = header_font
        ws.cell(5, i).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.cell(5, i).border = thin
    ws.row_dimensions[5].height = 30

    last_data = 5 + len(ROWS)
    for r, row in enumerate(ROWS, 6):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
            ws.cell(r, c).font = body
            ws.cell(r, c).alignment = wrap
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = cream if r % 2 == 0 else white_f
        ws.row_dimensions[r].height = 50
        ws.cell(r, 1).alignment = center
        ws.cell(r, 4).fill = status_fill(str(row[3]))
        ws.cell(r, 4).font = bold
        ws.cell(r, 4).alignment = center

    sum_row = last_data + 2
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)
    ws.cell(sum_row, 1, "Frontend summary for manager")
    ws.cell(sum_row, 1).font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws.cell(sum_row, 1).fill = header_fill
    ws.cell(sum_row, 1).alignment = center

    summary = [
        ["Done on frontend", "All 17 table items: login + MFA, roles, isolation, server-only account create/update/offboard, audit log, usage (own org), validation, safe logs, session revoke, Dependabot, HTTPS headers, notification bell."],
        ["Partial", "None remaining on the frontend table. App Check / IP rate limit still sit on Firebase / backend."],
        ["Open (ops / console)", "Enable Identity Platform TOTP MFA in Firebase Console. Deploy rules + functions + Vercel."],
        ["Deploy still needed", "Firestore rules + Cloud Functions + Vercel redeploy + VITE_USE_MOCK=false. Until deploy, isolation is in code but not live."],
    ]
    for i, (k, v) in enumerate(summary, sum_row + 1):
        ws.cell(i, 1, k).font = bold
        ws.cell(i, 1).fill = surface
        ws.cell(i, 1).alignment = wrap
        ws.cell(i, 1).border = thin
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        ws.cell(i, 2, v).font = body
        ws.cell(i, 2).alignment = wrap
        ws.cell(i, 2).border = thin
        ws.row_dimensions[i].height = 42
        for c in range(2, 6):
            ws.cell(i, c).border = thin

    legend = sum_row + 6
    ws.merge_cells(start_row=legend, start_column=1, end_row=legend, end_column=5)
    ws.cell(legend, 1, "Colour: Green = Done   |   Amber = Partial   |   Red = Open")
    ws.cell(legend, 1).font = sub_font

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 52
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:E{last_data}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.oddFooter.center.text = "Voicera Frontend FINAL | HL-SEC-DEVBRIEF-2026-001 | Confidential"


def add_manager_sheet(wb: Workbook):
    ws = wb.create_sheet("Manager Summary")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Voicera Frontend — Manager Summary (FINAL)"
    ws["A1"].font = title_font

    ws.merge_cells("A2:D2")
    ws["A2"] = f"HL-SEC-DEVBRIEF-2026-001  |  {date.today().strftime('%d %b %Y')}  |  Confidential — Internal"
    ws["A2"].font = sub_font

    counts = {"Done": 0, "Partial": 0, "Open": 0}
    for row in ROWS:
        counts[str(row[3])] = counts.get(str(row[3]), 0) + 1

    ws["A4"] = "Scorecard"
    ws["A4"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws["A4"].fill = header_fill
    ws.merge_cells("A4:C4")
    for c in range(1, 4):
        ws.cell(4, c).fill = header_fill
        ws.cell(4, c).border = thin

    for i, (label, fill) in enumerate([("Done", done_f), ("Partial", part_f), ("Open", open_f)], 5):
        ws.cell(i, 1, label).font = bold
        ws.cell(i, 1).fill = fill
        ws.cell(i, 1).alignment = center
        ws.cell(i, 1).border = thin
        ws.cell(i, 2, counts[label]).font = bold
        ws.cell(i, 2).alignment = center
        ws.cell(i, 2).border = thin
        ws.cell(i, 2).fill = fill
        ws.cell(i, 3, f"{round(counts[label] / len(ROWS) * 100)}%").font = body
        ws.cell(i, 3).alignment = center
        ws.cell(i, 3).border = thin

    ws["A9"] = "Deploy checklist (ops — not frontend code)"
    ws["A9"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws["A9"].fill = header_fill
    ws.merge_cells("A9:D9")
    for c in range(1, 5):
        ws.cell(9, c).fill = header_fill
        ws.cell(9, c).border = thin

    for i, h in enumerate(["#", "Action", "Owner", "Status"], 1):
        ws.cell(10, i, h)
        ws.cell(10, i).font = header_font
        ws.cell(10, i).fill = header_fill
        ws.cell(10, i).alignment = center
        ws.cell(10, i).border = thin

    deploy = [
        [1, "Deploy Firestore rules (tenant isolation + audit_events)", "Ops", "Open"],
        [2, "Deploy Cloud Functions (create/update/offboard, audit, session revoke)", "Ops", "Open"],
        [3, "Vercel redeploy with VITE_USE_MOCK=false + security headers", "Ops", "Open"],
        [4, "Enable Identity Platform TOTP MFA in Firebase Console", "Ops", "Open"],
    ]
    for r, row in enumerate(deploy, 11):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
            ws.cell(r, c).font = body
            ws.cell(r, c).alignment = wrap
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = cream if r % 2 == 0 else white_f
        ws.row_dimensions[r].height = 36
        ws.cell(r, 1).alignment = center
        ws.cell(r, 4).fill = status_fill(str(row[3]))
        ws.cell(r, 4).font = bold
        ws.cell(r, 4).alignment = center

    ws["A16"] = "Bottom line"
    ws["A16"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws["A16"].fill = header_fill
    ws.merge_cells("A16:D16")
    for c in range(1, 5):
        ws.cell(16, c).fill = header_fill
        ws.cell(16, c).border = thin

    ws.merge_cells("A17:D17")
    ws["A17"] = (
        "Frontend work for HL-SEC-DEVBRIEF is complete in code (17/17 Done). "
        "Do not call production signed-off until the 4 deploy checklist items above are done."
    )
    ws["A17"].font = bold
    ws["A17"].alignment = wrap
    ws["A17"].fill = surface
    ws.row_dimensions[17].height = 48
    for c in range(1, 5):
        ws.cell(17, c).border = thin

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.center.text = "Voicera Frontend FINAL | Confidential"


def build():
    wb = Workbook()
    add_table_sheet(wb)
    add_manager_sheet(wb)
    out = r"c:\karan\voicera\Voicera_Frontend_Security_Governance_FINAL.xlsx"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    build()
